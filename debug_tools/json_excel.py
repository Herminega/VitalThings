import json
import os
import pandas as pd

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo

from collections import defaultdict
from datetime import datetime, timedelta
from dateutil import parser

# Data file
input_file = 'random/split_file_254.jsonl' #'devicelog_preprocessing/devicelog_processed.json'

### ---------------------------------------------- ###
### ---- DECLEARING VAR, DICT, LISTS and FUNC ---- ###
### ---------------------------------------------- ###

# How long we evaluate the data
days_ago = 90
max_timestamp_date = (datetime.now() - timedelta(days=days_ago)).date()

# Limits deciding if unit is sick
voltage_level = 4.7
rssi_level = -85
min_number_of_lowVoltage = 2
max_number_of_lowVoltage = 15
min_number_of_lowRssi = 3
max_number_of_lowRssi = 15
min_number_of_commReset = 30
max_number_of_commReset = 50
min_number_of_boot = 30
max_number_of_boot = 50
min_number_of_fatalError = 3
max_number_of_fatalError = 5
min_number_of_XeThruStartup = 15 #If state: starting, sucsess: false
max_number_of_XeThruStartup = 30 #If state: starting, sucsess: false
min_number_of_uptime = 0
max_number_of_uptime = 0
# Initialize a dictionary to store consecutive flagged days for each unit ID
consecutive_flagged_days = defaultdict(int)
# Define the threshold for consecutive flagged days
threshold_consecutive_flagged = 5

# Empty lists, will contain sick and flagged units, and their data
unique_sick_units = set()
unique_flagged_units = set()
sick_units = []
flagged_units = []
data=[]

# Initialize a dictionary to store dates for each unit ID
dates_by_unit = defaultdict(list)
# Create a dictionary to keep track of unit categorization
unit_categorization = {}
# Initialize a dictionary to store daily event frequencies for each unit ID
daily_type_frequencies = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
daily_resetreason_frequencies = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
daily_lowvoltage_frequencies = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
daily_lowrssi_frequencies = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
# Initialize a low rssi/voltage list to calculate average
low_voltage_values = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))
low_rssi_values = defaultdict(lambda: defaultdict(lambda: defaultdict(int)))

# Limits deciding if cell should be colord or not
threshold_yellow_commReset = 10  # commReset
threshold_red_commReset = 20    # commReset
threshold_yellow_XeThruStartup = 10 # xethrustartup
threshold_red_XeThruStartup = 20    # xethrustartup
threshold_yellow_GuidedBreathing = 10 #guidedbreathing
threshold_red_GuidedBreathing = 20 #guidedbreathing
threshold_yellow_boot= 10  # boot
threshold_red_boot = 20    # boot
threshold_yellow_fatal = 2  # fatal
threshold_red_fatal = 4    # fatal
threshold_yellow_bootloader = 10  # bootloader
threshold_red_bootloader = 20    # bootloader
threshold_yellow_watchdog= 10  # watchdog
threshold_red_watchdog = 20    # watchdog
threshold_yellow_alarm = 20 # alarm
threshold_red_alarm = 30 # alarm
threshold_yellow_cleared = 20 # cleared
threshold_red_cleared = 30 # cleared
threshold_yellow_softReset = 20 #softReset
threshold_red_softReset = 30 #softReset
threshold_yellow_voltage= 7  # voltage
threshold_red_voltage = 10    # voltage
threshold_yellow_rssi= 5  # rssi
threshold_red_rssi = 8    # rssi

### -------------------------------------- ###
### ---- MANGING DATA FROM INPUT FILE ---- ###
### -------------------------------------- ###

# Reading data
with open(input_file, 'r') as file:
    for line in file:
        json_obj = json.loads(line)
        unit_id = json_obj['UnitID']
        timestamp_str = json_obj['Timestamp']
        timestamp = parser.parse(timestamp_str)
        timestamp_date = timestamp.date()

        # Check if timestamp_date for unitID is less than "days_ago"
        if timestamp_date < max_timestamp_date:
            continue
        
        dates_by_unit.setdefault(unit_id, []).append(timestamp_date.strftime('%Y-%m-%d'))   

        type_value = json_obj.get('Type', json_obj.get('ContentType', ''))
        daily_type_frequencies[unit_id][timestamp_date][type_value] += 1

        reset_reason = json_obj.get('Payload', {}).get('ResetReason', '') or json_obj.get('Content', {}).get('ResetReason', 'No Reset Reason')
        daily_resetreason_frequencies[unit_id][timestamp_date][reset_reason] += 1
        
        voltage = json_obj.get('Payload', {}).get('Voltage', '') or json_obj.get('Content', {}).get('ResetReason', '')
        if voltage and voltage < voltage_level:
            daily_lowvoltage_frequencies[unit_id][timestamp_date]['Low voltage'] += 1
            low_voltage_values[unit_id][timestamp_date]['Sum voltage'] += voltage

        rssi = json_obj.get('Payload', {}).get('NetworkInfo', {}).get('Rssi', '') or json_obj.get('Payload', {}).get('NetworkInfo', {}).get('Rssi', '')
        if rssi and rssi < rssi_level:
            daily_lowrssi_frequencies[unit_id][timestamp_date]['Low rssi'] += 1
            low_rssi_values[unit_id][timestamp_date]['Sum rssi'] += rssi

# Filtering data
def filter_flagged(unit_id, type_freq, reset_reason_freq, low_voltage_freq, low_rssi_freq):
    if (
    min_number_of_commReset < type_freq['CommReset'] < max_number_of_commReset or
    min_number_of_boot < type_freq['Boot'] < max_number_of_boot or

    (min_number_of_XeThruStartup < type_freq['XeThruStartup'] < max_number_of_XeThruStartup and
    json_obj.get('Payload', {}).get('Success', False) is False and
    json_obj.get('Payload', {}).get('State', '') == 'starting') or

    min_number_of_fatalError < reset_reason_freq['Fatal'] < max_number_of_fatalError or

    (any(value > min_number_of_lowVoltage and value < max_number_of_lowVoltage for value in low_voltage_freq.values())) or
    (any(value > min_number_of_lowRssi and value < max_number_of_lowRssi for value in low_rssi_freq.values()))):
        consecutive_flagged_days[unit_id] += 1  # Increment the consecutive flagged days counter
        return True 
    else:
        consecutive_flagged_days[unit_id] = 0  # Reset the consecutive flagged days counter
        return False

def filter_sick(unit_id, type_freq, reset_reason_freq, low_voltage_freq, low_rssi_freq):
    if (
    # If a type, reset reason, rssi or voltage is "wrong"
    (type_freq['CommReset'] > max_number_of_commReset or
    type_freq['Boot'] > max_number_of_boot or

    (type_freq['XeThruStartup'] > max_number_of_XeThruStartup and
    json_obj.get('Payload', {}).get('Success', False) is False and
    json_obj.get('Payload', {}).get('State', '') == 'starting') or

    reset_reason_freq['Fatal'] > max_number_of_fatalError or

    (any(value > max_number_of_lowVoltage for value in low_voltage_freq.values())) or
    (any(value > max_number_of_lowRssi for value in low_rssi_freq.values())))   
    
    # If unit has been flagged for more than threshold_consecutive_flagged
    or (filter_flagged(unit_id, type_freq, reset_reason_freq, low_voltage_freq, low_rssi_freq) and consecutive_flagged_days[unit_id] >= threshold_consecutive_flagged)):
        #print(f'{consecutive_flagged_days}, {threshold_consecutive_flagged}')
        consecutive_flagged_days[unit_id] = 0  # Reset the consecutive flagged days counter
        return True
    else:
        consecutive_flagged_days[unit_id] = 0  # Reset the consecutive flagged days counter
        return False

# Formatig data
def format_data(event_freq, reset_freq, voltage_freq, rssi_freq):
    type_freq_str = '\n'.join([f"{event_type}: {count}" for event_type, count in event_freq.items()])
    reset_reason_freq_str = '\n'.join([f"{reason}: {count}" for reason, count in reset_freq.items()])
    low_voltage_freq_str = '\n'.join([f"{reason}: {count}" for reason, count in voltage_freq.items()])
    low_rssi_freq_str = '\n'.join([f"{reason}: {count}" for reason, count in rssi_freq.items()])

    if 'Low voltage' in voltage_freq:
        sum_voltage = low_voltage_values[unit_id][date]['Sum voltage']
        num_low_voltage = daily_lowvoltage_frequencies[unit_id][date]['Low voltage']
        if num_low_voltage > 0:
            average_voltage = sum_voltage / num_low_voltage
            low_voltage_values[unit_id][date]['Average voltage'] = average_voltage
            average_voltage_str = f"Average voltage: {average_voltage:.2f}"
            low_voltage_freq_str += f"\n{average_voltage_str}"

    if 'Low rssi' in rssi_freq:
        sum_rssi = low_rssi_values[unit_id][date]['Sum rssi']
        num_low_rssi = daily_lowrssi_frequencies[unit_id][date]['Low rssi']
        if num_low_rssi > 0:
            average_rssi = sum_rssi / num_low_rssi
            low_rssi_values[unit_id][date]['Average rssi'] = average_rssi
            average_rssi_str = f"Average rssi: {average_rssi:.2f}"
            low_rssi_freq_str += f"\n{average_rssi_str}"


    return type_freq_str, reset_reason_freq_str, low_voltage_freq_str, low_rssi_freq_str

# Appending data to data_list
def append_data(data_list, unit_id, date, type_freq_str, resetreason_freq_str, lowvoltage_freq_str, lowrssi_freq_str):
    data_list.append({
        'UnitID': unit_id,
        'Date': date.strftime('%Y-%m-%d'),
        'Number of type': type_freq_str,
        'Number of reset reason': resetreason_freq_str,
        'Number of low voltage': lowvoltage_freq_str,
        'Number of low rssi': lowrssi_freq_str
    })

# Create dataframe by reading and formatting data in directories 
for unit_id, date_type_freq in daily_type_frequencies.items():
    # Initialize the consecutive_flagged_days for this unit
    consecutive_flagged_days[unit_id] = 0
    # Initialize unit categorization if it doesn't exist
    if unit_id not in unit_categorization:
        unit_categorization[unit_id] = {
            'is_sick': False,
            'is_flagged': False,
        }

    for date, type_freq in date_type_freq.items(): 
        # Number of reset reason, voltage, and rssi for each unitID per date
        reset_reason_freq = daily_resetreason_frequencies[unit_id][date]
        low_voltage_freq = daily_lowvoltage_frequencies[unit_id][date]
        low_rssi_freq = daily_lowrssi_frequencies[unit_id][date]

        # Formatting data
        type_freq_str, resetreason_freq_str, lowvoltage_freq_str, lowrssi_freq_str = format_data(type_freq, reset_reason_freq, low_voltage_freq, low_rssi_freq)

        # Get status for unitID per date
        is_sick = filter_sick(unit_id, type_freq, reset_reason_freq, low_voltage_freq, low_rssi_freq)
        is_flagged = filter_flagged(unit_id, type_freq, reset_reason_freq, low_voltage_freq, low_rssi_freq)
        
        # Update unit categorization
        unit_categorization[unit_id]['is_sick'] = unit_categorization[unit_id]['is_sick'] or is_sick
        unit_categorization[unit_id]['is_flagged'] = unit_categorization[unit_id]['is_flagged'] or is_flagged 

        # Append data only if unit is marked as 'sick' or 'flagged'
        if is_sick or is_flagged:
            append_data(data, unit_id, date, type_freq_str, resetreason_freq_str, lowvoltage_freq_str, lowrssi_freq_str)

# Create lists based on unit categorization
for unit_id, categorization in unit_categorization.items():
    if categorization['is_sick']:
        if unit_id not in unique_sick_units:
            unique_sick_units.add(unit_id)
            sick_units.append({'UnitID': unit_id})
        
    elif categorization['is_flagged']:
        if unit_id not in unique_flagged_units:
            unique_flagged_units.add(unit_id)
            flagged_units.append({'UnitID': unit_id})

# Creating DataFrames from lists
df_flagged_units = pd.DataFrame(flagged_units)
df_sick_units = pd.DataFrame(sick_units)
df_data = pd.DataFrame(data)

### --------------------------------------------------------- ###
### ---- CREATING EXCEL FILE WITH FILTERING AND COLORING ---- ###
### --------------------------------------------------------- ###

# Create a new workbook
wb = Workbook()
# Delete default sheet
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# Add the flagged DataFrame to the workbook, as a worksheet
ws_flagged_units = wb.create_sheet(title="Flagged units")
for row in dataframe_to_rows(df_flagged_units, index=False, header=True):
    ws_flagged_units.append(row)

ws_sick_units = wb.create_sheet(title="Sick units")
for row in dataframe_to_rows(df_sick_units, index=False, header=True):
    ws_sick_units.append(row)

ws_data = wb.create_sheet(title="Data- units")
for row in dataframe_to_rows(df_data, index=False, header=True):
    ws_data.append(row)

# Set font for the header row
header_font = Font(bold=True)  # You can customize other font properties as well
# Set font for header row in both worksheets
for ws in [ws_data, ws_flagged_units, ws_sick_units]:
    for cell in ws[1]:
        cell.font = header_font

# Add filtering and color on UNITID-text to both worksheets
ws_data.auto_filter.ref = ws_data.dimensions
prev_unit_id = None  # Initialize previous unit ID
col_idx = 1  # Column index for UNITID-text
for row in ws_data.iter_rows(min_row=2, max_row=ws_data.max_row, min_col=col_idx, max_col=col_idx):
    for cell in row:
        # Apply font color based on prev_unit_id
        if prev_unit_id == cell.value:
            cell.font= Font(color='808080', sz=10)
        else:
            cell.font = Font(color='000000')

        # Update prev_unit_id
        prev_unit_id = cell.value

# Define the thresholds for different event types
thresholds = {
    'CommReset': {'yellow': threshold_yellow_commReset, 'red': threshold_red_commReset},
    'XeThruStartup': {'yellow': threshold_yellow_XeThruStartup, 'red': threshold_red_XeThruStartup},
    'GuidedBreathing': {'yellow': threshold_yellow_GuidedBreathing, 'red': threshold_red_GuidedBreathing},
    'Boot': {'yellow': threshold_yellow_boot, 'red': threshold_red_boot},
    'Fatal': {'yellow': threshold_yellow_fatal, 'red': threshold_red_fatal},
    'Bootloader': {'yellow': threshold_yellow_bootloader, 'red': threshold_red_bootloader},
    'Watchdog': {'yellow': threshold_yellow_watchdog, 'red': threshold_red_watchdog},
    'Alarm': {'yellow': threshold_yellow_alarm, 'red': threshold_red_alarm},
    'Cleared': {'yellow': threshold_yellow_cleared, 'red': threshold_red_cleared},
    'SoftReset': {'yellow': threshold_yellow_softReset, 'red': threshold_red_softReset},
    'Low voltage': {'yellow': threshold_yellow_voltage, 'red': threshold_red_voltage},
    'Low rssi': {'yellow': threshold_yellow_rssi, 'red': threshold_red_rssi}
}

# Create a function to apply cell color based on conditions
def apply_cell_color(value, thresholds):
    # If there is no value
    if value is None:
        return None, None
    
    # Read the value, and split on line shift
    parts = value.split('\n')
    event_counts = {}

    # Read every line within the parts (cell)
    for part in parts:
        event_name_and_count = part.split(':')
        if len(event_name_and_count) == 2:
            event_name, count_str = event_name_and_count
            if event_name.strip() not in ["Ping", "No Reset Reason", "Average rssi", "Average voltage"]:
                numeric_value = int(count_str.strip())
                event_counts[event_name.strip()] = numeric_value

    max_event_type = None
    max_numeric_value = 0

    for event_name, count in event_counts.items():
        if count >= thresholds[event_name]['red']:
            max_event_type = event_name
            max_numeric_value = count
            break
        elif count >= thresholds[event_name]['yellow'] and not max_event_type:
            max_event_type = event_name
            max_numeric_value = count

    if max_event_type:
        if max_numeric_value >= thresholds[max_event_type]['red']:
            return PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"), Border(left=Side(style='thin', color='C0C0C0'), right=Side(style='thin', color='C0C0C0'), top=Side(style='thin', color='C0C0C0'), bottom=Side(style='thin', color='C0C0C0'))
        elif max_numeric_value >= thresholds[max_event_type]['yellow']:
            return PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"), Border(left=Side(style='thin', color='C0C0C0'), right=Side(style='thin', color='C0C0C0'), top=Side(style='thin', color='C0C0C0'), bottom=Side(style='thin', color='C0C0C0'))

    return None, None  # Return None if no color is applied

columns_to_color = ['Number of type', 'Number of reset reason', 'Number of low voltage', 'Number of low rssi']
# Color cells based on your thresholds
for col_idx, col in enumerate(columns_to_color, start=3):
    col_letter = get_column_letter(col_idx)
    for row in ws_data.iter_rows(min_row=2, max_row=ws_data.max_row, min_col=col_idx, max_col=col_idx):
        for cell in row:
            new_fill, border_style = apply_cell_color(cell.value, thresholds)
            if new_fill is not None:
                cell.fill = new_fill
            cell.border = border_style

# Save the workbook and write to excel
excel_file_path = 'output_with_sheets.xlsx'

# Re-size cells with regards to the text
for sheet in wb:
    for column_cells in sheet.columns:
        max_length = 0
        column = column_cells[0].column_letter  # Get the column name
        
        for cell in column_cells:
            cell.alignment = Alignment(wrapText = True)
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width_small = (max_length)  # Add some padding
        adjusted_width_large = (max_length + 2)*1.2  # Add some padding 

        sheet.column_dimensions[column].width = adjusted_width_large # Column A and B 

        for col in ['C', 'D', 'E', 'F']:
            sheet.column_dimensions[col].width = adjusted_width_small # Column C, D, E and F
        
wb.save(excel_file_path)
print('Dataframe saved to excel')
  
os.system(f'open "{excel_file_path}"')
